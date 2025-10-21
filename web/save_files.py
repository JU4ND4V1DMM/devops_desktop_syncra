import os
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import polars as pl
import pandas as pd
from typing import Union

def format_excel_file(filepath: str):
    """
    Formats the header of a single Excel file (XLSX) using openpyxl by applying bold, 
    center alignment, and freezing the first row. This function is file I/O based.
    """
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        ws.freeze_panes = ws['A2']

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(filepath)
        print(f"✨ Formatted header in file: {filepath}")
    except Exception as e:
        print(f"ERROR: Could not format Excel file at {filepath}. Reason: {e}")

def save_to_csv(data_frame, output_path: str, filename: str, partitions: Union[int, str], delimiter: str = ","):
    """
    Saves a DataFrame (Polars or PySpark) to a single CSV file.
    Tries PySpark first for performance, falls back to Polars if conversion fails.
    """
    partitions = int(partitions)
    if output_path and not os.path.exists(output_path):
        os.makedirs(output_path)
    
    now = datetime.now()
    time_file = now.strftime("%Y%m%d_%H%M")
    file_date = now.strftime("%Y%m%d")
    
    # Create a temporary folder to store initial files
    temp_folder_name = f"{filename}_{time_file}"
    temp_output_path = os.path.join(output_path, temp_folder_name)

    # Save the DataFrame into the temporary folder using PySpark
    try:
        (data_frame
            .repartition(partitions) 
            .write
            .mode("overwrite")
            .option("header", "true")
            .option("delimiter", delimiter)
            .csv(temp_output_path)
        )

        # Remove unnecessary files
        for root, dirs, files in os.walk(temp_output_path):
            for file in files:
                if file.startswith("._") or file == "_SUCCESS" or file.endswith(".crc"):
                    os.remove(os.path.join(root, file))

        # Move the CSV files from the temporary folder to the main output path
        for i, file in enumerate(os.listdir(temp_output_path), start=1):
            if file.endswith(".csv"):
                old_file_path = os.path.join(temp_output_path, file)
                new_file_path = os.path.join(output_path, f"{filename} {file_date} {i}.csv")
                os.rename(old_file_path, new_file_path)

        # Delete the temporary folder
        if os.path.exists(temp_output_path):
            shutil.rmtree(temp_output_path)
            print(f"🗑️ Temporary folder deleted: {temp_output_path}")

        print(f"✔️ CSV files successfully moved to: {output_path}")

    except Exception as e:
        print(f"⚠️ PySpark CSV save failed ({e}), trying Polars fallback...")
        try:
            # Save using Polars
            print("💾 Attempting to save with Polars...")
            if not isinstance(data_frame, pl.DataFrame):
                # Convert from PySpark or Pandas
                if hasattr(data_frame, "toPandas"):  # PySpark
                    data_frame = pl.from_pandas(data_frame.toPandas())
                elif isinstance(data_frame, pd.DataFrame):
                    data_frame = pl.from_pandas(data_frame)

            data_frame.write_csv(file=os.path.join(output_path, f"{filename} {file_date}.csv"), separator=delimiter, include_header=True)
            print(f"✅ CSV file successfully saved to: {output_path}")

        except Exception as spark_err:
            print(f"❌ ERROR: Failed to save with both PySpark and Polars. Reason: {spark_err}")

def save_to_0csv(data_frame, output_path: str, filename: str, partitions: Union[int, str], delimiter: str = ","):
    """
    Saves a DataFrame (Polars or PySpark) to a '.0csv' file.
    Tries PySpark first, falls back to Polars if needed.
    """
    partitions = int(partitions)
    if output_path and not os.path.exists(output_path):
        os.makedirs(output_path)
    
    now = datetime.now()
    time_file = now.strftime("%Y%m%d_%H%M")
    file_date = now.strftime("%Y%m%d")
    
    # Create a temporary folder to store initial files
    temp_folder_name = f"{filename}_{time_file}"
    temp_output_path = os.path.join(output_path, temp_folder_name)

    # Save the DataFrame into the temporary folder using PySpark
    try:
        (data_frame
            .repartition(partitions)
            .write
            .mode("overwrite")
            .option("header", "true")
            .option("delimiter", delimiter)
            .csv(temp_output_path)
        )

        # Remove unnecessary files
        for root, dirs, files in os.walk(temp_output_path):
            for file in files:
                if file.startswith("._") or file == "_SUCCESS" or file.endswith(".crc"):
                    os.remove(os.path.join(root, file))

        # Move the CSV files from the temporary folder to the main output path
        for i, file in enumerate(os.listdir(temp_output_path), start=1):
            if file.endswith(".csv"):
                old_file_path = os.path.join(temp_output_path, file)
                new_file_path = os.path.join(output_path, f"{filename} {file_date} {i}.0csv")
                os.rename(old_file_path, new_file_path)

        # Delete the temporary folder
        if os.path.exists(temp_output_path):
            shutil.rmtree(temp_output_path)
            print(f"🗑️ Temporary folder deleted: {temp_output_path}")

        print(f"✔️ .0csv files successfully moved to: {output_path}")

    except Exception as e:
        print(f"⚠️ PySpark .0csv save failed ({e}), trying Polars fallback...")
        try:
            # Save using Polars
            print("💾 Attempting to save with Polars (.0csv)...")
            if not isinstance(data_frame, pl.DataFrame):
                if hasattr(data_frame, "toPandas"):  # PySpark
                    data_frame = pl.from_pandas(data_frame.toPandas())
                elif isinstance(data_frame, pd.DataFrame):
                    data_frame = pl.from_pandas(data_frame)

            data_frame.write_csv(file=os.path.join(output_path, f"{filename} {file_date}.0csv"), separator=delimiter, include_header=True)
            print(f"✅ .0csv file successfully saved to: {output_path}")

        except Exception as spark_err:
            print(f"❌ ERROR: Failed to save with both PySpark and Polars. Reason: {spark_err}")

def save_to_xlsx(data_frame: pl.DataFrame, output_path: str, filename: str, partitions: Union[int, str]):
    """
    Saves a Polars DataFrame to a single Excel (XLSX) file via Pandas, 
    then applies custom openpyxl formatting.
    """
    if output_path and not os.path.exists(output_path):
        os.makedirs(output_path)

    now = datetime.now()
    file_date = now.strftime("%Y%m%d")
    folder_name = f"{filename} {file_date}"

    full_output_path = os.path.join(output_path, folder_name, filename)

    print(f"💾 Saving DataFrame as Excel to {full_output_path}...")
    try:
        # Convert Polars or PySpark to Pandas
        if hasattr(data_frame, "toPandas"):  # PySpark
            df_pandas = data_frame.toPandas()
        elif isinstance(data_frame, pl.DataFrame):
            df_pandas = data_frame.to_pandas()
        else:
            df_pandas = data_frame

        df_pandas.to_excel(excel_writer=full_output_path, sheet_name='Details', index=False)
        format_excel_file(full_output_path)
        print(f"✅ Final Excel file saved as: {full_output_path}")

    except Exception as e:
        print(f"❌ ERROR: Failed to save DataFrame to XLSX. Reason: {e}")
