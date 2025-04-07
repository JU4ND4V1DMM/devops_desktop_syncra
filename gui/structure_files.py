import os
import csv
from openpyxl import load_workbook
from datetime import datetime

def details_files(folder_path, output_folder):
    # Create a valid output file path
    folder_name = os.path.basename(folder_path)
    output_file = os.path.join(output_folder, f"Validación Archivos {folder_name} {datetime.now().strftime('%Y-%m-%d')}.csv")
    
    # Open the output file for writing
    with open(output_file, mode='w', newline='', encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file, delimiter=';')
        writer.writerow(['Archivo', 'Nombre Hoja', 'Titulos'])
        
        # Iterate through the files in the folder
        for filename in os.listdir(folder_path):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(folder_path, filename)
                workbook = load_workbook(file_path, read_only=True)
                sheet_names = workbook.sheetnames
                
                # Write titles from each sheet
                for sheet_name in sheet_names:
                    sheet = workbook[sheet_name]
                    titles = [cell.value for cell in sheet[1]]
                    writer.writerow([filename, sheet_name, ', '.join(map(str, titles))])